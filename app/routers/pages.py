# app/routers/pages.py
from datetime import datetime, timezone, timedelta
from html import escape
from io import BytesIO
from pathlib import Path

from fastapi import APIRouter, Request
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

router = APIRouter()

@router.get("/", response_class=HTMLResponse)
def welcome(request: Request):
    templates = request.app.state.templates
    cfg = request.app.state.cfg
    return templates.TemplateResponse(
        "welcome.html",
        {"request": request, "cfg": cfg, "page": "welcome"},
    )

@router.get("/game", response_class=HTMLResponse)
def game(request: Request):
    templates = request.app.state.templates
    cfg = request.app.state.cfg
    player = request.session.get("player")
    if not player:
        return RedirectResponse(url="/", status_code=303)
    return templates.TemplateResponse(
        "game.html",
        {"request": request, "cfg": cfg, "player": player, "page": "game"},
    )

@router.get("/scoreboard", response_class=HTMLResponse)
def scoreboard(request: Request):
    templates = request.app.state.templates
    cfg = request.app.state.cfg
    player = request.session.get("player") or {"name": "Player"}
    score = request.session.get("last_score", 0)

    def to_ordinal(n: int) -> str:
        if 10 <= n % 100 <= 20:
            suffix = "th"
        else:
            suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
        return f"{n}{suffix}"

    score_service = request.app.state.scores
    top_entries = score_service.top()
    decorated_top = [
        {
            "id": entry.id,
            "rank": entry.rank,
            "ordinal": to_ordinal(entry.rank),
            "name": entry.name,
            "score": entry.score,
            "created_at": entry.created_at,
            "is_player": False,
        }
        for entry in top_entries
    ]

    last_score_id_raw = request.session.get("last_score_id")
    try:
        last_score_id = int(last_score_id_raw) if last_score_id_raw is not None else None
    except (TypeError, ValueError):
        last_score_id = None
    player_entry = score_service.fetch(last_score_id) if last_score_id else None
    if player_entry:
        player_rank = {
            "id": player_entry.id,
            "rank": player_entry.rank,
            "ordinal": to_ordinal(player_entry.rank),
            "name": player_entry.name,
            "score": player_entry.score,
            "created_at": player_entry.created_at,
            "is_player": True,
        }
        if player_entry.rank > score_service.top_n:
            decorated_top.append({"ellipsis": True})
            decorated_top.append(player_rank)
        else:
            for row in decorated_top:
                if row.get("id") == player_entry.id:
                    row["is_player"] = True
                    break
    else:
        player_rank = None

    return templates.TemplateResponse(
        "scoreboard.html",
        {
            "request": request,
            "cfg": cfg,
            "player": player,
            "score": score,
            "rows": decorated_top,
            "player_entry": player_entry,
            "player_rank": player_rank,
            "page": "scoreboard",
        },
    )


@router.get("/admin", response_class=HTMLResponse)
def admin_panel(request: Request):
    templates = request.app.state.templates
    base = Path(__file__).resolve().parents[2]
    with (base / "config.json").open("r", encoding="utf-8") as f:
        cfg_raw = f.read()
    return templates.TemplateResponse(
        "admin.html",
        {"request": request, "cfg_raw": cfg_raw, "page": "admin"},
    )


@router.get("/admin/export")
def admin_export(request: Request):
    score_service = request.app.state.scores
    entries = score_service.all_with_rank()
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")

    wb = Workbook()
    ws = wb.active
    ws.title = "Leaderboard"

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="222222")
    header_align = Alignment(horizontal="center")

    headers = ["Rank", "Name", "Phone", "Score", "Recorded At (UTC)"]
    col_widths = [8, 30, 20, 12, 25]

    for col, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = width

    for row, entry in enumerate(entries, start=2):
        ws.cell(row=row, column=1, value=entry.rank)
        ws.cell(row=row, column=2, value=entry.name)
        ws.cell(row=row, column=3, value=entry.phone or "")
        ws.cell(row=row, column=4, value=entry.score)
        ws.cell(row=row, column=5, value=entry.created_at + " WIB")

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    headers_resp = {"Content-Disposition": f"attachment; filename=leaderboard-{timestamp}.xlsx"}
    return Response(
        content=buffer.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_resp,
    )
